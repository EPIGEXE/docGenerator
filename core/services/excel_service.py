import polars as pl
from pathlib import Path
from typing import Dict, List, Optional
from dataclasses import dataclass
import os
from openpyxl import Workbook
import re

@dataclass
class ExcelTemplate:
    columns: List[str]
    sample_data: List[Dict]
    data_types: Dict[str, str]
    
class ExcelAnalyzer:
    def __init__(self):
        self.supported_extensions = {'.xlsx', '.xls', '.csv'}

    async def analyze_template(self, file_path: str | Path) -> ExcelTemplate:
        """Excel 템플릿 파일 분석"""
        try:
            file_path = Path(file_path)
            if file_path.suffix not in self.supported_extensions:
                raise ValueError(f"지원하지 않는 파일 형식입니다: {file_path.suffix}")

            # Polars를 사용하여 데이터 로드
            if file_path.suffix == '.csv':
                df = pl.read_csv(file_path)
            else:
                df = pl.read_excel(file_path)

            # 컬럼 정보 추출
            columns = df.columns

            # 데이터 타입 분석
            data_types = {
                col: self._infer_column_type(df[col])
                for col in columns
            }

            # 샘플 데이터 추출 (최대 5개 행)
            sample_data = df.head(5).to_dicts()

            return ExcelTemplate(
                columns=columns,
                sample_data=sample_data,
                data_types=data_types
            )

        except Exception as e:
            raise Exception(f"템플릿 분석 중 오류 발생: {str(e)}")

    def _infer_column_type(self, column: pl.Series) -> str:
        """컬럼 데이터 타입 추론"""
        if column.dtype in [pl.Int8, pl.Int16, pl.Int32, pl.Int64]:
            return 'number'
        elif column.dtype in [pl.Float32, pl.Float64]:
            return 'decimal'
        elif column.dtype == pl.Date:
            return 'date'
        elif column.dtype == pl.Datetime:
            return 'datetime'
        elif column.dtype == pl.Boolean:
            return 'boolean'
        return 'text'

class ExcelService:
    def __init__(self):
        self.analyzer = ExcelAnalyzer()

    def _sanitize_sheet_name(self, name: str) -> str:
        """Excel 시트 이름에 사용할 수 없는 문자 제거 및 길이 제한"""
        # 허용되지 않는 문자 제거
        safe_name = re.sub(r'[\[\]:*?/\\]', '_', name)
        # 시트 이름 길이 제한 (31자)
        return safe_name[:31]

    async def generate_excel(self, analyzed_content: dict) -> str:
        """Excel 파일 생성"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"  # 기본 시트 이름 설정
            
            # 데이터가 있는지 확인
            if not analyzed_content or 'data' not in analyzed_content:
                raise ValueError("데이터가 없거나 잘못된 형식입니다")
                
            data = analyzed_content['data']
            if not data:
                raise ValueError("데이터가 비어있습니다")
                
            # 헤더 추가 (첫 번째 데이터의 키를 사용)
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 데이터 추가
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
            
            # 열 너비 자동 조정
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # 파일 저장
            output_path = os.path.join('media', 'documents', 'temp.xlsx')
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            
            return output_path
                
        except Exception as e:
            raise Exception(f"Excel 생성 중 오류 발생: {str(e)}")

    async def analyze_and_generate(self, template_path: str | Path, analyzed_content: dict) -> dict:
        """템플릿 분석 및 컨텐츠 생성"""
        try:
            # 템플릿 분석
            template = await self.analyzer.analyze_template(template_path)
            
            # AI 서비스에 전달할 템플릿 정보 구성
            template_info = {
                "columns": template.columns,
                "sample_data": template.sample_data,
                "data_types": template.data_types,
                "patterns": self._extract_patterns(template)
            }
            
            return template_info

        except Exception as e:
            raise Exception(f"템플릿 분석 중 오류 발생: {str(e)}")

    def _extract_patterns(self, template: ExcelTemplate) -> Dict:
        """데이터 패턴 추출"""
        patterns = {}
        
        for col in template.columns:
            sample_values = [row.get(col) for row in template.sample_data if row.get(col)]
            if sample_values:
                patterns[col] = {
                    "unique_values": list(set(sample_values)),
                    "value_length": {
                        "min": min(len(str(v)) for v in sample_values),
                        "max": max(len(str(v)) for v in sample_values)
                    }
                }
        
        return patterns